__author__ = 'liangl2'
# https://openpyxl.readthedocs.org/en/latest/tutorial.html#loading-from-a-file
from openpyxl import Workbook
from openpyxl import load_workbook
from hxg import diversity_inclusion_search

wb = load_workbook(filename="websites audit.xlsx")
ws = wb.active
url_list = []
row_num = 1
for cell in ws.columns[1]:  # get column B value cell by cell, ws.column[1][:].value
    print("Processing "+str(row_num)+'/'+str(len(ws.columns[0]))+': '\
          +ws.cell(row=row_num, column=1).value+"...")

    if cell.value is not None and type(cell.value) is not bool and \
                    len(cell.value) > 3 and "http" in cell.value:
        url = cell.value
        result = diversity_inclusion_search(url)
        has_DI_index = 'C'+str(row_num)
        has_search_index = 'E'+str(row_num)

        ws[has_DI_index] = result['has_DI']*"Yes"+(not result['has_DI'])*"No"
        ws[has_search_index] = result['has_search']*"Yes"+(not result['has_search'])*"No"
        wb.save("websites audit.xlsx")

    row_num += 1