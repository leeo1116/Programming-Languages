__author__ = 'liangl2'
# https://openpyxl.readthedocs.org/en/latest/tutorial.html#loading-from-a-file
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename="websites  audit.xlsx")
ws = wb.active
url_list = []
for cell in ws.columns[1]:  # get column B value cell by cell, ws.column[1][:].value
    if cell.value is not None and len(cell.value) > 3 and "http" in cell.value:
        url_list.append(cell.value)
# url_list = ws["B2":"B1190"].value
print('\n'.join(url_list[:100]))





"""
print(ws.cell(row=1, column=1).value)
print(wb.get_sheet_names())

import datetime

wb = Workbook()
ws = wb.active
ws["A1"] = 16
ws.append([11, 2, 3])
ws["A3"] = datetime.datetime.now()
print(ws.cell(row=2, column=2).value)
wb.save("excel_example.xlsx")
"""