__author__ = 'liangl2'


def main():
    from openpyxl import Workbook
    from linecache import getline
    import re
    wb = Workbook()
    ws = wb.active
    ws.title = "Eye Diagram Variation"

    n_serdes = 4
    n_index = 10
    BER = 1e-10
    atten = 0
    pre = 0
    post = 0
    for serdes in range(n_serdes):
        for index in range(n_index):

            # Generate tap-coefficients of TX FFE
            r = serdes*n_index+index+1
            ws.cell(row = r, column = 1).value = atten
            ws.cell(row = r, column = 2).value = pre
            ws.cell(row = r, column = 3).value = post

            # Read RX CTLE, RX DFE settings and eye height/width
            filename = "simba_data\\9 Inch SerDes0X06 Eye Diagram Variation\\"\
                       +"0x"+"{s:02d}".format(s = serdes+6)+"_000_"+"{i:02d}".format(i = index+1)+".ebert"
            print("Processing file "+filename)

            eye_width_line_num = find_line_num(filename, "Eye width at "+str(BER))
            eye_height_line_num = find_line_num(filename, "Eye height at "+str(BER))
            CTLE_DC_line = getline(filename, 55)
            CTLE_LF_line = getline(filename, 56)
            CTLE_HF_line = getline(filename, 57)
            CTLE_BW_line = getline(filename, 58)
            DFE_gain_line = getline(filename, 66)
            DFE_taps_line = getline(filename, 61)
            eye_height_line = getline(filename, eye_height_line_num)
            eye_width_line = getline(filename, eye_width_line_num)

            pattern = r"[-+]?\d*\.\d+|[-+]?\d+"
            CTLE_DC_str = re.findall(pattern, CTLE_DC_line)
            CTLE_LF_str = re.findall(pattern, CTLE_LF_line)
            CTLE_HF_str = re.findall(pattern, CTLE_HF_line)
            CTLE_BW_str = re.findall(pattern, CTLE_BW_line)
            DFE_gain_str = re.findall(pattern, DFE_gain_line)
            DFE_taps_str = re.findall(pattern, DFE_taps_line)
            eye_height_str = re.findall(r"(\d+)(?:\s)?mV", eye_height_line)
            eye_width_str = re.findall(r"(\d+)(?:\s)?mUI", eye_width_line)

            # Write RX CTLE, RX DFE settings and eye height/width to excel sheet
            ws.cell(row = r, column = 4).value = int(CTLE_DC_str[0])
            ws.cell(row = r, column = 5).value = int(CTLE_LF_str[0])
            ws.cell(row = r, column = 6).value = int(CTLE_HF_str[0])
            ws.cell(row = r, column = 7).value = int(CTLE_BW_str[0])
            ws.cell(row = r, column = 8).value = int(DFE_gain_str[0])
            ws.cell(row = r, column = 9).value = ', '.join(DFE_taps_str[0:5])
            ws.cell(row = r, column = 10).value = ', '.join(DFE_taps_str[5:12])
            ws.cell(row = r, column = 11).value = int(eye_height_str[0])
            ws.cell(row = r, column = 12).value = int(eye_width_str[0])

    wb.save("simba_eye_diagram_variation.xlsx")


def find_line_num(filename, pattern):
    with open(filename, 'r') as f:
        lines_list = f.readlines()
    for i, line in enumerate(lines_list):
        if pattern in line:
            return i+1

if __name__ == '__main__':
    main()